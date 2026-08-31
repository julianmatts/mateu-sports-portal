# -*- coding: utf-8 -*-
"""
Genera el Excel de Uniformes para el supervisor.

Lee la dotacion real del ultimo periodo de data/indicadores/ (una persona por
cada quien registro ventas en su sucursal) y arma un libro con:
  1. Instructivo
  2. Talles              -> relevamiento: una fila por persona, talle por prenda
  3. Entregas            -> registro de lo entregado (fecha, prenda, talle, cant.)
  4. Resumen produccion  -> unidades de cada prenda x talle a pedir
  5. Resumen por sucursal-> avance del relevamiento y unidades entregadas
  6. Listas              -> hoja oculta con los valores de las validaciones

Uso:  python scripts/gen-uniformes.py [periodo] [salida.xlsx]
"""
import json, os, sys, glob

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATOS = os.path.join(RAIZ, 'data', 'indicadores')

NAVY = '0B1527'
ROJO = 'CC0000'
GRIS = 'F5F7FC'
BORDE_GRIS = 'D8DEE9'

PRENDAS = ['Remera', 'Campera', 'Pantalon']
TALLES_LETRA = ['XS', 'S', 'M', 'L', 'XL', 'XXL', '3XL']
TALLES_NUM = [str(t) for t in range(36, 60, 2)]

# grupo del ETL -> puesto que se entiende en la sucursal
PUESTO = {
    'Jefatura': 'Encargado/a',
    'Ventas': 'Vendedor/a',
    'Caja': 'Cajero/a',
    'Refuerzos': 'Deposito / refuerzo',
    'Gerencia / otros': 'Gerencia',
    'Otros': 'Otros',
}

# sucursales sin datos de venta en Indicadores: van igual en las listas y en el
# resumen, con las filas en blanco para que carguen la gente a mano
SUC_EXTRA = ['10-Diagonal 80']

# "vendedores" del sistema que no son personas (canales de venta)
NO_PERSONAS = {'MERCADOLIBRE', 'TIENDANUBE', 'MOSTRADOR', 'SIN VENDEDOR', 'VENTA WEB'}


def es_persona(nombre):
    n = (nombre or '').strip().upper()
    return bool(n) and n not in NO_PERSONAS and not n.startswith('WEB ')


def ultimo_periodo():
    ps = sorted(d for d in os.listdir(DATOS)
                if os.path.isdir(os.path.join(DATOS, d)))
    return ps[-1] if ps else None


def regimen(sector):
    s = (sector or '').lower()
    if 'full' in s:
        return 'Full Time'
    if 'part' in s:
        return 'Part Time'
    return ''


def leer_dotacion(periodo):
    """Una fila por persona de cada sucursal, sin coberturas."""
    filas = []
    for ruta in sorted(glob.glob(os.path.join(DATOS, periodo, '*.json'))):
        if os.path.basename(ruta) == 'cadena.json':
            continue
        with open(ruta, encoding='utf-8') as f:
            d = json.load(f)
        suc = d.get('sucursal') or os.path.basename(ruta)[:-5]
        for v in d.get('vendedores', []):
            if v.get('cubre'):   # cobertura: la persona ya figura en su sucursal fija
                continue
            if not es_persona(v.get('vendedor')):
                continue
            filas.append({
                'sucursal': suc,
                'persona': (v.get('vendedor') or '').strip(),
                'puesto': PUESTO.get(v.get('grupo'), v.get('grupo') or ''),
                'regimen': regimen(v.get('sector')),
            })
    vistos, unicas = set(), []
    for f in filas:
        k = (f['sucursal'], f['persona'])
        if k in vistos or not f['persona']:
            continue
        vistos.add(k)
        unicas.append(f)
    unicas.sort(key=lambda f: (f['sucursal'], f['persona']))
    return unicas


# ---------- helpers de estilo ----------
def encabezado(ws, fila, titulos, anchos=None):
    fill = PatternFill('solid', fgColor=NAVY)
    fuente = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    borde = Border(bottom=Side('thick', color=ROJO))
    for i, t in enumerate(titulos, start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.fill, c.font, c.border = fill, fuente, borde
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[fila].height = 30
    if anchos:
        for i, a in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = a


def titulo_hoja(ws, texto, sub=''):
    ws['A1'] = 'MATEU SPORTS'
    ws['A1'].font = Font(name='Calibri', bold=True, size=10, color=ROJO)
    ws['A2'] = texto
    ws['A2'].font = Font(name='Calibri', bold=True, size=16, color=NAVY)
    if sub:
        ws['A3'] = sub
        ws['A3'].font = Font(name='Calibri', size=10, color='55627A')


def bordear(ws, r1, c1, r2, c2):
    lado = Side('thin', color=BORDE_GRIS)
    b = Border(left=lado, right=lado, top=lado, bottom=lado)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = b


def main():
    periodo = sys.argv[1] if len(sys.argv) > 1 else ultimo_periodo()
    salida = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.expanduser('~'), 'Downloads', 'Uniformes Sucursales.xlsx')

    dot = leer_dotacion(periodo)
    sucursales = sorted({f['sucursal'] for f in dot} | set(SUC_EXTRA))
    personas = sorted({f['persona'] for f in dot})
    n = len(dot)
    print('Periodo %s: %d personas en %d sucursales' % (periodo, n, len(sucursales)))

    wb = Workbook()

    # ---------------- 1. Instructivo ----------------
    ws = wb.active
    ws.title = 'Instructivo'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 110
    titulo_hoja(ws, 'Uniformes de sucursal',
                'Relevamiento de talles y control de entregas - dotacion al periodo %s' % periodo)
    lineas = [
        ('t', 'Como se completa'),
        ('p', '1. Hoja TALLES: ya viene con toda la gente cargada (sucursal, nombre y puesto). Cada sucursal completa el talle de remera, campera y pantalon de cada persona. Los talles se eligen del desplegable de la celda.'),
        ('p', '2. Si falta alguien (ingreso nuevo), se agrega al final de la hoja Talles escribiendo la sucursal y el nombre; los desplegables siguen funcionando en las filas vacias.'),
        ('p', '3. Si una persona ya no trabaja, escribir BAJA en Observaciones. No borrar la fila: asi queda el historial.'),
        ('t', 'Entregas'),
        ('p', '4. Hoja ENTREGAS: una fila por cada entrega (fecha, sucursal, persona, prenda, talle y cantidad). Sirve para saber quien ya recibio y quien no.'),
        ('t', 'Que se calcula solo'),
        ('p', 'RESUMEN PRODUCCION: cuantas unidades de cada prenda y talle hay que pedir, segun lo cargado en Talles, y cuanto falta entregar. Es la hoja que se le pasa al proveedor.'),
        ('p', 'RESUMEN POR SUCURSAL: cuanta gente tiene cada sucursal, cuantos talles faltan cargar y cuantas unidades se entregaron.'),
        ('t', 'Aclaraciones'),
        ('p', 'Remera y campera usan talles de letra (XS a 3XL). Pantalon usa talle numerico (36 a 58).'),
        ('p', 'La dotacion sale del modulo Mi Sucursal (Indicadores) del portal: son las personas que registraron ventas en el periodo %s. No incluye coberturas (cada persona figura en su sucursal fija).' % periodo),
        ('p', 'Para regenerar el archivo con la dotacion actualizada: python scripts/gen-uniformes.py'),
    ]
    r = 5
    for tipo, txt in lineas:
        c = ws.cell(row=r, column=2, value=txt)
        if tipo == 't':
            c.font = Font(bold=True, size=12, color=NAVY)
        else:
            c.font = Font(size=11)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[r].height = 32
        r += 1
    ws.sheet_view.showGridLines = False

    # ---------------- Listas (oculta) ----------------
    wl = wb.create_sheet('Listas')
    wl['A1'] = 'Talles letra'
    wl['B1'] = 'Talles numero'
    wl['C1'] = 'Prendas'
    wl['D1'] = 'Sucursales'
    wl['E1'] = 'Personas'
    for i, v in enumerate(TALLES_LETRA, start=2):
        wl.cell(row=i, column=1, value=v)
    for i, v in enumerate(TALLES_NUM, start=2):
        wl.cell(row=i, column=2, value=v)
    for i, v in enumerate(PRENDAS, start=2):
        wl.cell(row=i, column=3, value=v)
    for i, v in enumerate(sucursales, start=2):
        wl.cell(row=i, column=4, value=v)
    for i, v in enumerate(personas, start=2):
        wl.cell(row=i, column=5, value=v)

    ref_letra = 'Listas!$A$2:$A$%d' % (len(TALLES_LETRA) + 1)
    ref_num = 'Listas!$B$2:$B$%d' % (len(TALLES_NUM) + 1)
    ref_prenda = 'Listas!$C$2:$C$%d' % (len(PRENDAS) + 1)
    ref_suc = 'Listas!$D$2:$D$%d' % (len(sucursales) + 1)
    ref_per = 'Listas!$E$2:$E$%d' % (len(personas) + 1)

    # ---------------- 2. Talles ----------------
    wt = wb.create_sheet('Talles', 1)
    titulo_hoja(wt, 'Relevamiento de talles',
                'Completar el talle de cada persona. Remera y campera: XS a 3XL. Pantalon: 36 a 58.')
    FIL0 = 5
    encabezado(wt, FIL0,
               ['Sucursal', 'Apellido y nombre', 'Puesto', 'Regimen',
                'Talle remera', 'Talle campera', 'Talle pantalon', 'Observaciones'],
               [26, 32, 20, 12, 14, 14, 15, 34])
    fill_alt = PatternFill('solid', fgColor=GRIS)
    for i, f in enumerate(dot):
        r = FIL0 + 1 + i
        wt.cell(row=r, column=1, value=f['sucursal'])
        wt.cell(row=r, column=2, value=f['persona'])
        wt.cell(row=r, column=3, value=f['puesto'])
        wt.cell(row=r, column=4, value=f['regimen'])
        for c in range(5, 8):
            wt.cell(row=r, column=c).alignment = Alignment(horizontal='center')
        if i % 2:
            for c in range(1, 9):
                wt.cell(row=r, column=c).fill = fill_alt
    ULT = FIL0 + max(n, 1) + 60   # filas extra para altas nuevas
    bordear(wt, FIL0 + 1, 1, ULT, 8)
    dv_l = DataValidation(type='list', formula1='=%s' % ref_letra, allow_blank=True)
    dv_n = DataValidation(type='list', formula1='=%s' % ref_num, allow_blank=True)
    wt.add_data_validation(dv_l)
    wt.add_data_validation(dv_n)
    dv_l.add('E%d:F%d' % (FIL0 + 1, ULT))
    dv_n.add('G%d:G%d' % (FIL0 + 1, ULT))
    wt.auto_filter.ref = 'A%d:H%d' % (FIL0, ULT)
    wt.freeze_panes = 'C%d' % (FIL0 + 1)
    wt.sheet_view.showGridLines = False

    # ---------------- 3. Entregas ----------------
    we = wb.create_sheet('Entregas', 2)
    titulo_hoja(we, 'Registro de entregas',
                'Una fila por entrega. La cantidad suele ser 1; si se entregan 2 remeras iguales, poner 2.')
    E0 = 5
    encabezado(we, E0,
               ['Fecha', 'Sucursal', 'Apellido y nombre', 'Prenda', 'Talle',
                'Cantidad', 'Entregado por', 'Observaciones'],
               [12, 26, 32, 14, 10, 10, 22, 34])
    E_ULT = E0 + 500
    bordear(we, E0 + 1, 1, E_ULT, 8)
    for r in range(E0 + 1, E_ULT + 1):
        we.cell(row=r, column=1).number_format = 'dd/mm/yyyy'
        for c in (5, 6):
            we.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    dv_s = DataValidation(type='list', formula1='=%s' % ref_suc, allow_blank=True)
    dv_p = DataValidation(type='list', formula1='=%s' % ref_per, allow_blank=True)
    dv_pr = DataValidation(type='list', formula1='=%s' % ref_prenda, allow_blank=True)
    dv_t = DataValidation(type='list',
                          formula1='"%s"' % ','.join(TALLES_LETRA + TALLES_NUM),
                          allow_blank=True)
    for dv, col in ((dv_s, 'B'), (dv_p, 'C'), (dv_pr, 'D'), (dv_t, 'E')):
        we.add_data_validation(dv)
        dv.add('%s%d:%s%d' % (col, E0 + 1, col, E_ULT))
    we.auto_filter.ref = 'A%d:H%d' % (E0, E_ULT)
    we.freeze_panes = 'A%d' % (E0 + 1)
    we.sheet_view.showGridLines = False

    # ---------------- 4. Resumen produccion ----------------
    wp = wb.create_sheet('Resumen produccion', 3)
    titulo_hoja(wp, 'Resumen de produccion',
                'Unidades a pedir por prenda y talle, segun lo cargado en la hoja Talles.')
    rng_rem = 'Talles!$E$%d:$E$%d' % (FIL0 + 1, ULT)
    rng_cam = 'Talles!$F$%d:$F$%d' % (FIL0 + 1, ULT)
    rng_pan = 'Talles!$G$%d:$G$%d' % (FIL0 + 1, ULT)
    rng_ent_pr = 'Entregas!$D$%d:$D$%d' % (E0 + 1, E_ULT)
    rng_ent_ta = 'Entregas!$E$%d:$E$%d' % (E0 + 1, E_ULT)
    rng_ent_ca = 'Entregas!$F$%d:$F$%d' % (E0 + 1, E_ULT)

    P0 = 5
    encabezado(wp, P0, ['Talle', 'Remeras', 'Camperas', 'Remeras entregadas',
                        'Camperas entregadas', 'Falta remera', 'Falta campera'],
               [12, 12, 12, 19, 19, 14, 14])
    for i, t in enumerate(TALLES_LETRA):
        r = P0 + 1 + i
        wp.cell(row=r, column=1, value=t).font = Font(bold=True)
        wp.cell(row=r, column=2, value='=COUNTIF(%s,$A%d)' % (rng_rem, r))
        wp.cell(row=r, column=3, value='=COUNTIF(%s,$A%d)' % (rng_cam, r))
        wp.cell(row=r, column=4, value='=SUMIFS(%s,%s,"Remera",%s,$A%d)' % (rng_ent_ca, rng_ent_pr, rng_ent_ta, r))
        wp.cell(row=r, column=5, value='=SUMIFS(%s,%s,"Campera",%s,$A%d)' % (rng_ent_ca, rng_ent_pr, rng_ent_ta, r))
        wp.cell(row=r, column=6, value='=B%d-D%d' % (r, r))
        wp.cell(row=r, column=7, value='=C%d-E%d' % (r, r))
        for c in range(1, 8):
            wp.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    rt = P0 + 1 + len(TALLES_LETRA)
    wp.cell(row=rt, column=1, value='TOTAL').font = Font(bold=True, color=ROJO)
    for c in range(2, 8):
        cel = wp.cell(row=rt, column=c,
                      value='=SUM(%s%d:%s%d)' % (get_column_letter(c), P0 + 1,
                                                 get_column_letter(c), rt - 1))
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal='center')
    bordear(wp, P0 + 1, 1, rt, 7)

    Q0 = rt + 3
    encabezado(wp, Q0, ['Talle', 'Pantalones', '', 'Pantalones entregados', '',
                        'Falta pantalon', ''])
    for i, t in enumerate(TALLES_NUM):
        r = Q0 + 1 + i
        wp.cell(row=r, column=1, value=t).font = Font(bold=True)
        wp.cell(row=r, column=2, value='=COUNTIF(%s,$A%d)' % (rng_pan, r))
        wp.cell(row=r, column=4, value='=SUMIFS(%s,%s,"Pantalon",%s,$A%d)' % (rng_ent_ca, rng_ent_pr, rng_ent_ta, r))
        wp.cell(row=r, column=6, value='=B%d-D%d' % (r, r))
        for c in (1, 2, 4, 6):
            wp.cell(row=r, column=c).alignment = Alignment(horizontal='center')
    rt2 = Q0 + 1 + len(TALLES_NUM)
    wp.cell(row=rt2, column=1, value='TOTAL').font = Font(bold=True, color=ROJO)
    for c in (2, 4, 6):
        cel = wp.cell(row=rt2, column=c,
                      value='=SUM(%s%d:%s%d)' % (get_column_letter(c), Q0 + 1,
                                                 get_column_letter(c), rt2 - 1))
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal='center')
    bordear(wp, Q0 + 1, 1, rt2, 7)

    wp.cell(row=rt2 + 2, column=1,
            value='Personas sin talle cargado').font = Font(bold=True, color=NAVY)
    for j, (etq, rng) in enumerate((('Remera', rng_rem), ('Campera', rng_cam),
                                    ('Pantalon', rng_pan))):
        r = rt2 + 3 + j
        wp.cell(row=r, column=1, value=etq)
        wp.cell(row=r, column=2,
                value='=COUNTIFS(Talles!$B$%d:$B$%d,"<>",%s,"")' % (FIL0 + 1, ULT, rng))
        wp.cell(row=r, column=2).alignment = Alignment(horizontal='center')
    wp.sheet_view.showGridLines = False

    # ---------------- 5. Resumen por sucursal ----------------
    wr = wb.create_sheet('Resumen por sucursal', 4)
    titulo_hoja(wr, 'Avance por sucursal',
                'Cuanta gente tiene cada sucursal, cuanto falta relevar y cuanto se entrego.')
    R0 = 5
    encabezado(wr, R0, ['Sucursal', 'Personas', 'Talles cargados', 'Falta cargar',
                        '% relevado', 'Unidades a entregar', 'Entregadas', 'Pendientes'],
               [28, 11, 15, 13, 12, 19, 13, 13])
    rng_suc = 'Talles!$A$%d:$A$%d' % (FIL0 + 1, ULT)
    rng_nom = 'Talles!$B$%d:$B$%d' % (FIL0 + 1, ULT)
    rng_ent_suc = 'Entregas!$B$%d:$B$%d' % (E0 + 1, E_ULT)
    for i, s in enumerate(sucursales):
        r = R0 + 1 + i
        wr.cell(row=r, column=1, value=s)
        wr.cell(row=r, column=2, value='=COUNTIFS(%s,$A%d,%s,"<>")' % (rng_suc, r, rng_nom))
        wr.cell(row=r, column=3,
                value='=COUNTIFS(%s,$A%d,%s,"<>")+COUNTIFS(%s,$A%d,%s,"<>")+COUNTIFS(%s,$A%d,%s,"<>")'
                      % (rng_suc, r, rng_rem, rng_suc, r, rng_cam, rng_suc, r, rng_pan))
        wr.cell(row=r, column=4, value='=B%d*3-C%d' % (r, r))
        cel = wr.cell(row=r, column=5, value='=IF(B%d=0,"",C%d/(B%d*3))' % (r, r, r))
        cel.number_format = '0%'
        wr.cell(row=r, column=6, value='=C%d' % r)
        wr.cell(row=r, column=7, value='=SUMIFS(%s,%s,$A%d)' % (rng_ent_ca, rng_ent_suc, r))
        wr.cell(row=r, column=8, value='=F%d-G%d' % (r, r))
        for c in range(2, 9):
            wr.cell(row=r, column=c).alignment = Alignment(horizontal='center')
        if i % 2:
            for c in range(1, 9):
                wr.cell(row=r, column=c).fill = fill_alt
    rr = R0 + 1 + len(sucursales)
    wr.cell(row=rr, column=1, value='TOTAL').font = Font(bold=True, color=ROJO)
    for c in (2, 3, 4, 6, 7, 8):
        cel = wr.cell(row=rr, column=c,
                      value='=SUM(%s%d:%s%d)' % (get_column_letter(c), R0 + 1,
                                                 get_column_letter(c), rr - 1))
        cel.font = Font(bold=True)
        cel.alignment = Alignment(horizontal='center')
    cel = wr.cell(row=rr, column=5, value='=IF(B%d=0,"",C%d/(B%d*3))' % (rr, rr, rr))
    cel.font = Font(bold=True)
    cel.number_format = '0%'
    cel.alignment = Alignment(horizontal='center')
    bordear(wr, R0 + 1, 1, rr, 8)
    wr.sheet_view.showGridLines = False

    wl.sheet_state = 'hidden'
    wb.active = 1
    wb.save(salida)
    print('OK ->', salida)


if __name__ == '__main__':
    main()
